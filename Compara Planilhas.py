import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import os
import pyautogui
import time
from openpyxl import load_workbook

def log_message(message):
    text_log.config(state="normal")
    text_log.insert(tk.END, message + "\n")
    text_log.config(state="disabled")
    text_log.see(tk.END)

def geraPlanilha():
    pasta = filedialog.askdirectory(title="Selecione a pasta com os documentos")
    if not pasta:
        messagebox.showwarning("Aviso", "Nenhuma pasta selecionada.")
        return
    
    arquivos = [f for f in os.listdir(pasta) if not f.startswith('.')]
    nomes_documentos = [os.path.splitext(f)[0] for f in arquivos]
    formatos_documentos = [os.path.splitext(f)[1] for f in arquivos]
    
    df = pd.DataFrame({'Nome documento': nomes_documentos, 'Formato documento': formatos_documentos})
    df.to_excel('planilha_documentos.xlsx', index=False)
    log_message("Planilha gerada com sucesso!")
    messagebox.showinfo("Sucesso", "Planilha gerada com sucesso!")

def comparaPlanilhas():
    #xlsx_origem = filedialog.askopenfilename(title="Selecione a planilha para comparar", filetypes=[("Arquivos Excel", "*.xlsx")])
    #if not xlsx_origem:
    #    messagebox.showwarning("Aviso", "Nenhuma planilha selecionada.")
    #    return
    
    xlsx_origem = "lancamentos-mes.xlsx"
    
    xlsx_destino = "planilha_documentos.xlsx"
    df_origem = pd.read_excel(xlsx_origem, dtype=str)
    valores_origem = set(df_origem.iloc[:, 4].dropna().astype(str).str.strip())
    
    wb = load_workbook(xlsx_destino)
    ws = wb.active
    
    for row in range(2, ws.max_row + 1):
        valor_celula = str(ws[f"A{row}"].value).strip()
        ws[f"B{row}"] = "Encontrado" if valor_celula in valores_origem else "Não Encontrado"
    
    wb.save(xlsx_destino)
    log_message(f"Planilha '{xlsx_destino}' foi atualizada.")
    messagebox.showinfo("Sucesso", f"Planilha '{xlsx_destino}' foi atualizada.")

def gerar_relatorio_siat():

    data_inicio = entry_data_inicio.get()
    data_fim = entry_data_fim.get()

    if not data_inicio or not data_fim:
        messagebox.showwarning("Aviso", "Por favor, preencha as datas antes de gerar o relatório.")
        return

    # Passo 1: Abre o menu do Windows
    pyautogui.press("win")
    time.sleep(0.5)

    # Passo 2: Digita "SiatWEB"
    pyautogui.write("SiatWEB", interval=0)
    
    # Passo 2.5: Aguarda o Siat aparecer no menu pesquisar do Windows
    while True:
        pixel_color = pyautogui.pixel(64, 295) # Coordenadas do pixel monitorado
        if pixel_color == (52, 76, 143): # Cor azul escuro Siat
            break
            # Sai do loop quando o pixel assume a cor esperada
        time.sleep(0.1) # Aguarda 0.1 segundos antes de verificar novamente

    # Passo 3: Abre o aplicativo (pressionando Enter)
    pyautogui.press("enter")

    # Passo 4: Aguarda o SIAT abrir a tela de login
    while True:
        pixel_color = pyautogui.pixel(552, 377)  # Coordenadas do pixel monitorado
        if pixel_color == (156, 123, 37):  # Cor flicts (RGB)
            break  # Sai do loop quando o pixel assume a cor esperada
        time.sleep(0.1)  # Aguarda 0.1 segundos antes de verificar novamente

    # Passo 5: Digita a senha "4008"
    pyautogui.write("4008")

    # Passo 6: Pressiona "Tab" 4 vezes
    for _ in range(4):
        pyautogui.press("tab")

    # Passo 7: Aperta Ctrl + A (Seleciona tudo)
    pyautogui.hotkey("ctrl", "a")

    # Passo 8: Digita o usuário "ARTUR"
    pyautogui.write("ARTUR")

    # Passo 9: Aperta Enter
    pyautogui.press("enter")

    # **Passo intermediário: Aguarda o SIAT autenticar o login**
    while True:
        pixel_color = pyautogui.pixel(674, 383)  # Coordenadas do pixel monitorado
        if pixel_color == (229, 232, 241):  # Cor azul claro (RGB)
            break  # Sai do loop quando o pixel assume a cor esperada
        time.sleep(0.1)  # Aguarda 0.1 segundos antes de verificar novamente

    # Passo 10: Posiciona o mouse em X=93, Y=537
    pyautogui.moveTo(93, 537, duration=0)

    # Passo 11: Clica com o botão esquerdo (Frota)
    pyautogui.click()

    # Passo 11.5: Aguarda o menu "Frota" carregar
    while True:
        pixel_color = pyautogui.pixel(31, 125)  # Coordenadas do pixel monitorado
        if pixel_color == (0, 0, 0):  # Cor preta (RGB)
            break  # Sai do loop quando o pixel assume a cor esperada
        time.sleep(0.1)  # Aguarda 0.1 segundos antes de verificar novamente

    # Passo 12: Posiciona o mouse em X=80, Y=118
    pyautogui.moveTo(80, 118, duration=0)

    # Passo 13: Clica com o botão esquerdo (Documentos)
    pyautogui.click()

    # Passo 14: Posiciona o mouse em X=131, Y=145
    pyautogui.moveTo(131, 145, duration=0)

    # Passo 15: Clica com o botão esquerdo
    pyautogui.click()

    # **Passo 16: Aguarda a tela "Consulta - NF Entrada" carregar**
    while True:
        pixel_color = pyautogui.pixel(646, 130)  # Coordenadas do pixel monitorado
        if pixel_color == (58, 75, 135):  # Cor azul escuro (RGB)
            break  # Sai do loop quando o pixel assume a cor esperada
        time.sleep(0.1)  # Aguarda 0.1 segundos antes de verificar novamente

    # **Passo 17: Aperta Tab 17 vezes**
    for _ in range(17):
        pyautogui.press("tab")
        # time.sleep(0.1)

    # **Passo 18: Digita a primeira data**
    pyautogui.write(data_inicio, interval=0)

    # **Passo 19: Aperta Tab**
    pyautogui.press("tab")

    # **Passo 20: Digita a segunda data**
    pyautogui.write(data_fim, interval=0)

    # **Passo 21: Posiciona o mouse em X=428, Y=198 (CTG)**
    pyautogui.moveTo(428, 198, duration=0)

    # **Passo 22: Clica com o botão esquerdo**
    pyautogui.click()

    # **Passo 23: Posiciona o mouse em X=465, Y=429**
    pyautogui.moveTo(465, 429, duration=0.5)

    # **Passo 24: Clica com o botão esquerdo**
    pyautogui.click()

    # **Passo 25: Posiciona o mouse em X=1191, Y=125**
    pyautogui.moveTo(1191, 125, duration=0)

    # **Passo 26: Clique com o botão direito**
    pyautogui.click(button="right")

    # **Passo 27: Posiciona o mouse em X=1106, Y=203**
    pyautogui.moveTo(1106, 203, duration=0)

    # **Passo 28: Clique com o botão esquerdo**
    pyautogui.click()

    # **Passo 29: Posiciona o mouse em X=1334, Y=131**
    pyautogui.moveTo(1334, 131, duration=0)

    # **Passo 30: Clique com o botão esquerdo**
    pyautogui.click()

    # **Passo 31: Aguarda o Siat gerar o grid**
    while True:
        pixel_color = pyautogui.pixel(252, 574)  # Coordenadas do pixel a serem monitoradas
        if pixel_color == (255, 255, 255):  # Cor branca (RGB)
            break  # Sai do loop quando o pixel assume a cor esperada
        time.sleep(0.1)  # Aguarda 0.1 segundos antes de verificar novamente

    # **Passo 32: Posiciona o mouse em X=667, Y=571**
    pyautogui.moveTo(667, 571, duration=0)

    # **Passo 33: Clique com o botão direito**
    pyautogui.click(button="right")

    # **Passo 34: Posicione o mouse em X=746, Y=611**
    pyautogui.moveTo(746, 611, duration=0)

    # **Passo 35: Clique com o botão esquerdo**
    pyautogui.click()

    # Intervalo
    while True:
        pixel_color = pyautogui.pixel(255, 360)  # Coordenadas do pixel a serem monitoradas
        if pixel_color == (240, 240, 240):  # Cor cinza do explorador de arquivos (RGB)
            break  # Sai do loop quando o pixel assume a cor esperada
        time.sleep(0.1)  # Aguarda 0.1 segundos antes de verificar novamente

    # **Passo 36: Digita "lancamentos-mes"**
    pyautogui.write("lancamentos-mes", interval=0)

    # **Passo 37: Posiciona o mouse em X=1089, Y=46**
    # pyautogui.moveTo(1089, 46, duration=0.5)
    pyautogui.moveTo(1089, 46, duration=0)

    # **Passo 38: Clique com o botão esquerdo**
    pyautogui.click()

    # **Passo 39: Digita "Área de Trabalho"**
    pyautogui.write(r"C:\Users\Cliente\Desktop", interval=0)

    # **Passo 40: Aperta Enter**
    pyautogui.press("enter")

    # **Passo 41: Alt + L**
    pyautogui.hotkey("alt", "l")

    # **Passo 42: Enter 2 vezes**
    pyautogui.press("enter")
    time.sleep(0.5)
    pyautogui.press("enter")

    # **Passo 43: Abre o menu iniciar**
    pyautogui.press("win")

    # **Passo 44: Digita "lancamentos mes.csv"**
    pyautogui.write("lancamentos-mes.csv", interval=0.1)
    
    # **Passo 44.5: Aguarda o menu de pesquisa carregar**
    while True:
        pixel_color = pyautogui.pixel(68, 288)  # Coordenadas do pixel a serem monitoradas
        if pixel_color == (60, 188, 69):  # Cor verde LibreOffice Calc (RGB)
            break  # Sai do loop quando o pixel assume a cor esperada
        time.sleep(0.1)  # Aguarda 0.1 segundos antes de verificar novamente

    # **Passo 45: Aperta Enter para abrir o arquivo**
    pyautogui.press("enter")

    # **Passo 46: Aguarda o LibreOffice Calc abrir**
    while True:
        pixel_color = pyautogui.pixel(743, 117) # Coordenadas do pixel a serem monitoradas
        if pixel_color == (0, 120, 215): # Cor azul do botão OK (RGB)
            break # Sai do loop quando o pixel assume a cor esperada
        time.sleep(0.1) # Aguarda 0.1 segundos antes de verificar novamente

    # **Passo 47: Aperta Enter (caso haja algum aviso na abertura)**
    pyautogui.moveTo(891, 228, duration=0)
    pyautogui.click()
    pyautogui.press("enter")

    # **Passo 48: Aguarda a planilha abrir no LibreOffice Calc**
    while True:
        pixel_color = pyautogui.pixel(71, 60) # Coordenadas do pixel a serem monitoradas
        if pixel_color == (212, 146, 216): # Cor roxa do botão "Salvar" da planilha
            break # Sai do loop quando o pixel assume a cor esperada
        time.sleep(0.1) # Aguarda 0.1 segundos antes de verificar novamente
    

    # **Passo 49: Pressiona Ctrl + Shift + S (Salvar Como)**
    pyautogui.hotkey("ctrl", "shift", "s")
    time.sleep(1)

    # **Passo 50: Posiciona o mouse para a dropbox de extensões de arquivos**
    pyautogui.moveTo(370, 462, duration=0)

    # **Passo 51: Clique com botão esquerdo para abrir dropbox**
    pyautogui.click()

    # **Passo 52: Posiciona o mouse em cima do formato .xlsx (X=442, Y=582)**
    pyautogui.moveTo(302, 542, duration=1)

    # **Passo 53: Clique**
    pyautogui.click()

    # **Passo 54: Aperta Enter para salvar o arquivo .xlsx**
    pyautogui.press("enter")


    # **Passo 55: Alt + S para salvar**
    pyautogui.hotkey("alt", "s")

def deletar_documentos_repetidos():
    planilha_nome = "planilha_documentos.xlsx"
    pasta_arquivos = filedialog.askdirectory(title="Selecione a pasta com os arquivos")
    if not pasta_arquivos:
        messagebox.showwarning("Aviso", "Nenhuma pasta selecionada.")
        return
    
    try:
        df = pd.read_excel(planilha_nome, dtype=str)
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao carregar a planilha: {e}")
        return
    
    if "Nome documento" not in df.columns or "Formato documento" not in df.columns:
        messagebox.showerror("Erro", "A planilha não contém as colunas esperadas.")
        return
    
    arquivos_para_excluir = df[df["Formato documento"].str.strip().str.lower() == "encontrado"]["Nome documento"].str.strip()
    
    for arquivo in os.listdir(pasta_arquivos):
        nome_base, _ = os.path.splitext(arquivo)
        if nome_base in arquivos_para_excluir.values:
            caminho_arquivo = os.path.join(pasta_arquivos, arquivo)
            try:
                os.remove(caminho_arquivo)
                log_message(f"Arquivo deletado: {arquivo}")
            except Exception as e:
                log_message(f"Erro ao deletar {arquivo}: {e}")
    
    log_message("Processo concluído. Arquivos duplicados removidos.")
    messagebox.showinfo("Sucesso", "Processo concluído. Arquivos duplicados removidos.")

root = tk.Tk()
root.title("Gestor de Planilhas")
root.geometry("500x450")
root.resizable(False, False)

frame = ttk.Frame(root, padding=10)
frame.pack(fill="both", expand=True)

label = ttk.Label(frame, text="Escolha uma ação:", font=("Arial", 12, "bold"))
label.pack(pady=5)

frame_datas = ttk.Frame(frame)
frame_datas.pack(pady=5, fill="x")

label_data_inicio = ttk.Label(frame_datas, text="Data Início:")
label_data_inicio.pack(side="left", padx=5)
entry_data_inicio = ttk.Entry(frame_datas, width=10)
entry_data_inicio.pack(side="left")

label_data_fim = ttk.Label(frame_datas, text="Data Fim:")
label_data_fim.pack(side="left", padx=5)
entry_data_fim = ttk.Entry(frame_datas, width=10)
entry_data_fim.pack(side="left")

btn_gerar_relatorio = ttk.Button(frame_datas, text="Gerar Relatório de Lançamentos no SIAT", command=gerar_relatorio_siat)
btn_gerar_relatorio.pack(side="left", padx=5)

btn_gerar = ttk.Button(frame, text="Gerar Planilha", command=geraPlanilha)
btn_gerar.pack(pady=5, fill="x")

btn_comparar = ttk.Button(frame, text="Comparar Planilhas", command=comparaPlanilhas)
btn_comparar.pack(pady=5, fill="x")

btn_deletar = ttk.Button(frame, text="Deletar Notas Lançadas", command=deletar_documentos_repetidos)
btn_deletar.pack(pady=5, fill="x")

btn_sair = ttk.Button(frame, text="Sair", command=root.quit)
btn_sair.pack(pady=10, fill="x")

text_log = tk.Text(frame, height=8, wrap="word", state="normal")
text_log.pack(pady=5, fill="both", expand=True)
text_log.insert(tk.END, "Logs do sistema:\n")
text_log.config(state="disabled")

root.mainloop()
